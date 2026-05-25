"""
Weekly Experiment Logistics — Plan Generator + Input Template Builder
---------------------------------------------------------------------
Two entry points:
  1. generate_weekly_plan(week_data, output_path)  → formatted 3-sheet plan
  2. generate_input_template(output_path)           → blank weekly input template
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── palette ──────────────────────────────────────────────────────────────────
C_NAVY          = "1F3864"
C_RED           = "922B21"
C_PURPLE        = "6B5B8B"
C_ORANGE        = "F4B942"
C_YELLOW        = "FFF2CC"
C_SALMON        = "FCE4D6"
C_LAVENDER      = "D9D2E9"
C_LIGHT_GRAY    = "D9D9D9"
C_MID_GRAY      = "808080"
C_WHITE         = "FFFFFF"
C_SPACER        = "EBEBEB"

# template palette
C_TPL_HEADER    = "1F3864"
C_TPL_SECTION   = "2E4057"
C_TPL_INPUT     = "EBF5FB"
C_TPL_FIXED     = "F2F3F4"
C_TPL_WARN      = "FDEBD0"

# ── helpers ───────────────────────────────────────────────────────────────────
def _fill(c):  return PatternFill("solid", fgColor=c)
def _thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def _medium():
    s = Side(style="medium", color="555555")
    return Border(left=s, right=s, top=s, bottom=s)
def _thick_bottom():
    t = Side(style="medium", color="333333")
    n = Side(style="thin",   color="BBBBBB")
    return Border(left=n, right=n, top=n, bottom=t)
def _font(color="000000", sz=10, bold=False, name="Arial"):
    return Font(name=name, size=sz, bold=bold, color=color)
def _hfont(sz=10, bold=True):
    return Font(name="Arial", size=sz, bold=bold, color=C_WHITE)
def _al(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _s(cell, value=None, bg=None, font=None, align=None, border=None, num_format=None):
    if value is not None: cell.value = value
    if bg:     cell.fill   = _fill(bg)
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if border: cell.border = border
    if num_format: cell.number_format = num_format

def _merge(ws, r1, c1, r2, c2, value, bg, font, border=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    _s(cell, value, bg, font, align or _al(), border or _thin())

def _spacer_row(ws, row, n_cols, height=5):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row, c)
        cell.fill   = _fill(C_SPACER)
        cell.border = _thin()
    ws.row_dimensions[row].height = height

def _arrow(text):
    """Replace > with → in route strings."""
    return text.replace(">", "→") if text else ""


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1  –  STAFFING  (איושים)
# ══════════════════════════════════════════════════════════════════════════════
def _build_staffing(ws, wd):
    ws.sheet_view.rightToLeft  = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    experiments = wd["experiments"]

    # total column count: col A = role label, col B = loading, then day cols per experiment
    total_day_cols = sum(len(e["days"]) for e in experiments)
    total_cols = 2 + total_day_cols

    row = 1

    # ── week banner ────────────────────────────────────────────────────────
    _merge(ws, row, 1, row, total_cols,
           wd["week_label"], C_NAVY, _hfont(sz=13), _medium())
    ws.row_dimensions[row].height = 26
    row += 1

    # ── experiment name headers ────────────────────────────────────────────
    # col A and B: blank top-left corner
    for c in [1, 2]:
        cell = ws.cell(row, c)
        cell.fill   = _fill(C_NAVY)
        cell.border = _thin()

    col = 3
    for exp in experiments:
        n = len(exp["days"])
        _merge(ws, row, col, row, col + n - 1,
               exp["name"], C_NAVY, _hfont(sz=11), _medium())
        col += n

    ws.row_dimensions[row].height = 22
    row += 1

    # ── day sub-headers ────────────────────────────────────────────────────
    # role + loading header labels
    _s(ws.cell(row, 1), "תפקיד", C_ORANGE,
       _font("000000", 10, True), _al(), _thin())
    _s(ws.cell(row, 2), "העמסות", C_ORANGE,
       _font("000000", 10, True), _al(), _thin())

    col = 3
    for exp in experiments:
        for day in exp["days"]:
            _s(ws.cell(row, col), day, C_NAVY,
               _hfont(sz=9), _al(), _thin())
            col += 1

    ws.row_dimensions[row].height = 18
    row += 1

    # ── collect all unique role names (preserve order) ─────────────────────
    role_order = []
    for exp in experiments:
        for r in exp["roles"]:
            if r["role"] not in role_order:
                role_order.append(r["role"])

    # lookup: (exp_index, role_name) → role dict
    lookup = {}
    for ei, exp in enumerate(experiments):
        for r in exp["roles"]:
            key = (ei, r["role"])
            # support multiple rows of same role per experiment
            if key not in lookup:
                lookup[key] = []
            lookup[key].append(r)

    # find max number of rows any experiment uses for each role
    role_row_counts = {}
    for role_name in role_order:
        max_rows = 1
        for ei, exp in enumerate(experiments):
            rows_for_role = len(lookup.get((ei, role_name), []))
            max_rows = max(max_rows, rows_for_role)
        role_row_counts[role_name] = max_rows

    # ── role rows ──────────────────────────────────────────────────────────
    prev_role = None
    for role_name in role_order:
        n_rows = role_row_counts[role_name]

        for sub_row in range(n_rows):
            # role label (only on first sub-row, merged down if multiple)
            if sub_row == 0:
                if n_rows > 1:
                    _merge(ws, row, 1, row + n_rows - 1, 1,
                           role_name, C_ORANGE,
                           _font("000000", 10, True),
                           _thick_bottom(), _al("right", "center"))
                else:
                    _s(ws.cell(row, 1), role_name, C_ORANGE,
                       _font("000000", 10, True), _al("right"), _thick_bottom())

            # loading tick (first sub-row only, merged if needed)
            loading = False
            for ei in range(len(experiments)):
                rows = lookup.get((ei, role_name), [])
                if sub_row < len(rows) and rows[sub_row].get("loading"):
                    loading = True
            if sub_row == 0:
                tick = "☑" if loading else "☐"
                tick_bg = C_ORANGE if loading else C_LIGHT_GRAY
                if n_rows > 1:
                    _merge(ws, row, 2, row + n_rows - 1, 2,
                           tick, tick_bg, _font("000000", 11, loading),
                           _thick_bottom(), _al())
                else:
                    _s(ws.cell(row, 2), tick, tick_bg,
                       _font("000000", 11, loading), _al(), _thick_bottom())

            # day columns
            col = 3
            for ei, exp in enumerate(experiments):
                rows = lookup.get((ei, role_name), [])
                for day_i in range(len(exp["days"])):
                    cell = ws.cell(row, col)
                    val = ""
                    if sub_row < len(rows):
                        assigns = rows[sub_row].get("assignments", [])
                        if day_i < len(assigns):
                            val = assigns[day_i] or ""
                    _s(cell, val, C_YELLOW, _font(sz=10), _al(), _thick_bottom())
                    col += 1

            ws.row_dimensions[row].height = 17
            row += 1

        prev_role = role_name

    # ── totals row ─────────────────────────────────────────────────────────
    _s(ws.cell(row, 1), "סה\"כ", C_NAVY, _hfont(sz=10), _al(), _medium())
    _s(ws.cell(row, 2), "",     C_NAVY, None,            _al(), _medium())
    col = 3
    for exp in experiments:
        for i, total in enumerate(exp.get("totals", [0] * len(exp["days"]))):
            _s(ws.cell(row, col), total, C_NAVY, _hfont(sz=11), _al(), _medium())
            col += 1
    ws.row_dimensions[row].height = 20

    # ── column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 9
    col = 3
    for exp in experiments:
        for _ in exp["days"]:
            ws.column_dimensions[get_column_letter(col)].width = 24
            col += 1


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2  –  VEHICLES  (רכבים)
# ══════════════════════════════════════════════════════════════════════════════
def _build_vehicles(ws, wd):
    ws.sheet_view.rightToLeft  = True
    ws.sheet_view.showGridLines = False

    days     = wd["day_labels"]
    vehicles = wd["vehicles"]
    n_days   = len(days)
    # cols: A=vehicle name, B=detail label, then pairs (הלוך, חזור) per day
    total_cols = 2 + 2 * n_days

    row = 1

    # ── week banner ────────────────────────────────────────────────────────
    _merge(ws, row, 1, row, total_cols,
           wd["week_label"], C_RED, _hfont(sz=13), _medium())
    ws.row_dimensions[row].height = 26
    row += 1

    # ── day pair headers ───────────────────────────────────────────────────
    _s(ws.cell(row, 1), "רכבים",  C_RED, _hfont(), _al(), _medium())
    _s(ws.cell(row, 2), "פרטים",  C_RED, _hfont(), _al(), _thin())
    col = 3
    for day in days:
        _merge(ws, row, col, row, col + 1, day, C_RED, _hfont(), _medium())
        col += 2
    ws.row_dimensions[row].height = 20
    row += 1

    # ── הלוך / חזור sub-headers ────────────────────────────────────────────
    ws.cell(row, 1).fill   = _fill(C_RED)
    ws.cell(row, 1).border = _thin()
    ws.cell(row, 2).fill   = _fill(C_RED)
    ws.cell(row, 2).border = _thin()
    col = 3
    for _ in days:
        for label in ["הלוך", "חזור"]:
            _s(ws.cell(row, col), label, C_RED, _hfont(sz=9), _al(), _thin())
            col += 1
    ws.row_dimensions[row].height = 16
    row += 1

    # detail rows per vehicle (labels and their background colours)
    DETAILS = [
        ("מסלול",       C_ORANGE),
        ("שעות",        C_LIGHT_GRAY),
        ("מפקד רכב",    C_SALMON),
        ("נוסעים",      C_YELLOW),
        ("נוסעים",      C_YELLOW),
        ("הערות",       "F5F5F5"),
    ]

    for veh in vehicles:
        veh_start = row
        n_detail  = len(DETAILS)

        # vehicle name merged across all detail rows
        _merge(ws, row, 1, row + n_detail - 1, 1,
               veh["vehicle"],
               C_RED, _hfont(sz=10), _medium())

        for di, (label, bg) in enumerate(DETAILS):
            r = row + di

            # detail label
            _s(ws.cell(r, 2), label,
               "FCE4D6", _font("555555", 9, True), _al("right"), _thin())

            # day columns
            col = 3
            for day_data in veh.get("days", [None] * n_days):
                for direction in ["outbound", "return"]:
                    cell = ws.cell(r, col)
                    val  = ""
                    cbg  = bg

                    if day_data and day_data.get(direction):
                        d = day_data[direction]
                        if label == "מסלול":
                            val = _arrow(d.get("route", ""))
                            cbg = C_ORANGE
                        elif label == "שעות":
                            val = d.get("time", "")
                        elif label == "מפקד רכב":
                            val = d.get("commander", "")
                        elif label == "נוסעים":
                            pax = d.get("passengers", [])
                            idx = sum(1 for x in DETAILS[:di] if x[0] == "נוסעים")
                            val = pax[idx] if idx < len(pax) else ""
                        elif label == "הערות":
                            val = d.get("notes", "")
                            cbg = "F5F5F5"

                    _s(cell, val, cbg, _font(sz=10), _al(), _thin())
                    col += 1

            ws.row_dimensions[r].height = 17

        row += n_detail

        # spacer between vehicles — taller and clearly distinct
        for c in range(1, total_cols + 1):
            ws.cell(row, c).fill   = _fill(C_SPACER)
            ws.cell(row, c).border = _thin()
        ws.row_dimensions[row].height = 10
        row += 1

    # ── column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 13
    col = 3
    for _ in days:
        ws.column_dimensions[get_column_letter(col)].width     = 22
        ws.column_dimensions[get_column_letter(col+1)].width   = 22
        col += 2


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3  –  ACCOMMODATION  (לינה)
# ══════════════════════════════════════════════════════════════════════════════
def _build_rooms(ws, wd):
    ws.sheet_view.rightToLeft  = True
    ws.sheet_view.showGridLines = False

    nights   = wd["accommodation"]["nights"]
    n_nights = len(nights)
    # cols: A=unit, B=room, night cols, last = לא (not staying)
    total_cols = 2 + n_nights + 1

    row = 1

    # ── week banner ────────────────────────────────────────────────────────
    _merge(ws, row, 1, row, total_cols,
           wd["week_label"], C_PURPLE, _hfont(sz=13), _medium())
    ws.row_dimensions[row].height = 26
    row += 1

    # ── column headers ─────────────────────────────────────────────────────
    _s(ws.cell(row, 1), "צימרים", C_PURPLE, _hfont(), _al(), _medium())
    _s(ws.cell(row, 2), "חדרים",  C_PURPLE, _hfont(), _al(), _thin())
    col = 3
    for night in nights:
        _s(ws.cell(row, col), night, C_PURPLE, _hfont(), _al(), _thin())
        col += 1
    _s(ws.cell(row, col), "לא", C_PURPLE, _hfont(), _al(), _thin())
    ws.row_dimensions[row].height = 20
    row += 1

    for hostel in wd["accommodation"]["hostels"]:
        # hostel banner
        _merge(ws, row, 1, row, total_cols,
               hostel["name"], "404040", _hfont(sz=11), _medium())
        ws.row_dimensions[row].height = 18
        row += 1

        for unit in hostel["units"]:
            rooms   = unit["rooms"]
            n_rooms = len(rooms)

            # unit name merged down all rooms
            _merge(ws, row, 1, row + n_rooms - 1, 1,
                   unit["name"], C_LAVENDER,
                   _font("3D2566", 10, True), _medium())

            for rm in rooms:
                # room name
                _s(ws.cell(row, 2), rm["name"], C_LAVENDER,
                   _font("3D2566", 10, True), _al(), _thin())

                # night columns
                col = 3
                nights_data = rm.get("nights", [[]] * n_nights)
                for ni in range(n_nights):
                    occupants = nights_data[ni] if ni < len(nights_data) else []
                    val = "\n".join(occupants) if occupants else ""
                    _s(ws.cell(row, col), val, C_WHITE,
                       _font(sz=10), _al(), _thin())
                    col += 1

                # לא column (blank — filled manually or by agent)
                _s(ws.cell(row, col), "", "F9F9F9", _font(), _al(), _thin())

                h = max(17, len(max(nights_data or [[]], key=len, default=[])) * 14)
                ws.row_dimensions[row].height = h
                row += 1

        _spacer_row(ws, row, total_cols, 8)
        row += 1

    # ── totals row ─────────────────────────────────────────────────────────
    _s(ws.cell(row, 1), "סה\"כ:", C_LIGHT_GRAY,
       _font("000000", 10, True), _al("right"), _medium())
    _s(ws.cell(row, 2), "",      C_LIGHT_GRAY, None, _al(), _thin())
    col = 3
    totals = wd["accommodation"].get("night_totals", [0] * n_nights)
    for t in totals:
        _s(ws.cell(row, col), t, C_LIGHT_GRAY,
           _font("000000", 11, True), _al(), _thin())
        col += 1
    _s(ws.cell(row, col), "", C_LIGHT_GRAY, None, _al(), _thin())
    ws.row_dimensions[row].height = 20

    # ── column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    col = 3
    for _ in nights:
        ws.column_dimensions[get_column_letter(col)].width = 24
        col += 1
    ws.column_dimensions[get_column_letter(col)].width = 14


# ══════════════════════════════════════════════════════════════════════════════
# MAIN PLAN GENERATOR
# ══════════════════════════════════════════════════════════════════════════════
def generate_weekly_plan(week_data: dict, output_path: str):
    """
    Generate the 3-sheet weekly plan from a week_data dictionary.
    See INPUT TEMPLATE section below for the expected structure.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_staffing(wb.create_sheet("איושים"), week_data)
    _build_vehicles(wb.create_sheet("רכבים"),  week_data)
    _build_rooms(   wb.create_sheet("לינה"),   week_data)

    wb.save(output_path)
    print(f"✓ Plan saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# INPUT TEMPLATE GENERATOR
# ══════════════════════════════════════════════════════════════════════════════
def generate_input_template(output_path: str):
    """
    Generate a blank weekly input template that the team fills in each week.
    The agent reads this file and converts it into week_data for generate_weekly_plan().
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet A: WEEK SETUP ────────────────────────────────────────────────
    ws = wb.create_sheet("הגדרת שבוע")
    ws.sheet_view.rightToLeft  = True
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50

    def tpl_hdr(ws, row, text, bg=C_TPL_HEADER):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1)
        _s(c, text, bg, _hfont(sz=11), _al(), _medium())
        ws.row_dimensions[row].height = 22

    def tpl_row(ws, row, label, example="", note="", input_bg=C_TPL_INPUT):
        _s(ws.cell(row, 1), label,   C_TPL_FIXED, _font(sz=10, bold=True), _al("right"), _thin())
        _s(ws.cell(row, 2), example, input_bg,    _font(sz=10),            _al("right"), _thin())
        _s(ws.cell(row, 3), note,    "FAFAFA",    _font("888888", 9),      _al("left"),  _thin())
        ws.row_dimensions[row].height = 17

    r = 1
    _merge(ws, r, 1, r, 3, "📋  תבנית קלט שבועית — מלא לפני כל ניסוי",
           C_TPL_HEADER, _hfont(sz=13), _medium())
    ws.row_dimensions[r].height = 28
    r += 1

    _spacer_row(ws, r, 3, 6); r += 1

    tpl_hdr(ws, r, "פרטי שבוע"); r += 1
    tpl_row(ws, r, "תאריך התחלה", "12/05/26", "DD/MM/YY"); r += 1
    tpl_row(ws, r, "תאריך סיום",  "15/05/26", "DD/MM/YY"); r += 1
    tpl_row(ws, r, "שם השבוע",    "שבוע 3",  "לדוגמה: שבוע 3 / מבצע ירוק / ..."); r += 1

    _spacer_row(ws, r, 3, 6); r += 1

    tpl_hdr(ws, r, "ניסויים השבוע  (הוסף עמודה לכל ניסוי נוסף)"); r += 1
    # experiment columns header
    for ci, label in enumerate(["שדה", "ניסוי 1", "ניסוי 2", "ניסוי 3"], 1):
        bg = C_TPL_SECTION if ci == 1 else C_TPL_INPUT
        _s(ws.cell(r, ci), label, bg,
           _hfont() if ci == 1 else _font(sz=10, bold=True),
           _al(), _medium() if ci == 1 else _thin())
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35
    ws.row_dimensions[r].height = 18
    r += 1

    exp_fields = [
        ("אתר",            "עין יהב",     "מבוא חורון", ""),
        ("מנהל ניסוי",     "ספי",         "אריאל",      ""),
        ("יום התחלה",      "ראשון",       "חמישי",      ""),
        ("יום סיום",       "שלישי",       "חמישי",      ""),
        ("לינה? (כן/לא)",  "כן",          "לא",         ""),
        ("מלון / צימר",    "רגע בערבה",   "",           "רק אם יש לינה"),
    ]
    for label, ex1, ex2, note in exp_fields:
        _s(ws.cell(r, 1), label, C_TPL_FIXED, _font(sz=10, bold=True), _al("right"), _thin())
        _s(ws.cell(r, 2), ex1,   C_TPL_INPUT, _font(sz=10),            _al("right"), _thin())
        _s(ws.cell(r, 3), ex2,   C_TPL_INPUT, _font(sz=10),            _al("right"), _thin())
        _s(ws.cell(r, 4), "",    C_TPL_INPUT, _font(sz=10),            _al("right"), _thin())
        if note:
            _s(ws.cell(r, 5), f"← {note}", "FAFAFA", _font("888888", 9), _al("left"), _thin())
            ws.column_dimensions["F"].width = 30
        ws.row_dimensions[r].height = 17
        r += 1

    _spacer_row(ws, r, 3, 6); r += 1

    tpl_hdr(ws, r, "רכבים השבוע  (כן/לא לכל ניסוי)"); r += 1
    veh_header_cols = ["רכב", "ניסוי 1 — הלוך", "ניסוי 1 — חזור", "ניסוי 2 — הלוך", "ניסוי 2 — חזור", "מפקד רכב", "הערות"]
    for ci, label in enumerate(veh_header_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22
        bg = C_TPL_SECTION if ci == 1 else C_RED
        ft = _hfont(sz=9)
        _s(ws.cell(r, ci), label, bg, ft, _al(), _thin())
    ws.row_dimensions[r].height = 18
    r += 1

    sample_vehicles = [
        ("דוקאטו",                                      "כן", "כן", "", "", "שי ליסקובסקי", ""),
        ("טנדר #1 + מתדלקת סולר + נגרר ערבי",          "כן", "כן", "", "", "תומר",         "רשיון נגרר"),
        ("טנדר #2 + מתדלקת דס\"ל - גדול",              "כן", "כן", "", "", "יהב",           "רשיון נגרר"),
        ("יונדאי I20 1",                                 "",   "",  "כן","כן","", ""),
        ("יונדאי I20 2",                                 "",   "",  "כן","כן","", ""),
    ]
    for sv in sample_vehicles:
        for ci, val in enumerate(sv, 1):
            bg = C_TPL_FIXED if ci == 1 else (C_TPL_WARN if "נגרר" in str(sv[-1]) and ci == 7 else C_TPL_INPUT)
            _s(ws.cell(r, ci), val, bg, _font(sz=10), _al(), _thin())
        ws.row_dimensions[r].height = 17
        r += 1

    # ── Sheet B: STAFFING INPUT ────────────────────────────────────────────
    ws2 = wb.create_sheet("איושים — קלט")
    ws2.sheet_view.rightToLeft  = True
    ws2.sheet_view.showGridLines = False

    ROLES = [
        "מנהל ניסוי", "מטיס חוץ", "מטיס חוץ",
        "מטיס פנים", "מטיס פנים",
        "מהנדס בקרה", "מהנדס בקרה", "מהנדס מוביל", "מהנדס",
        "טכנאי מטוסים", "טכנאי מטוסים", "טכנאי מטוסים",
        "בטיחות",
        "אופרציה", "אופרציה", "אופרציה",
        "לוגיסטיקה", "לוגיסטיקה", "לוגיסטיקה",
        "מערך קרקעי", "מערך קרקעי",
    ]
    LOADING = {
        "טכנאי מטוסים": True, "בטיחות": True,
        "אופרציה": True, "לוגיסטיקה": True,
    }

    r2 = 1
    _merge(ws2, r2, 1, r2, 8,
           "איושים — מלא שם לכל תפקיד לכל יום  (ניסוי 1)",
           C_TPL_HEADER, _hfont(sz=12), _medium())
    ws2.row_dimensions[r2].height = 24
    r2 += 1

    # instruction
    _merge(ws2, r2, 1, r2, 8,
           "→  הוסף גיליון זהה לכל ניסוי נוסף השבוע",
           "F0F0F0", _font("666666", 9), _thin(), _al("right"))
    ws2.row_dimensions[r2].height = 15
    r2 += 1

    # header row
    day_cols = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי"]
    headers  = ["תפקיד", "העמסות"] + day_cols
    for ci, h in enumerate(headers, 1):
        bg = C_ORANGE if ci <= 2 else C_NAVY
        _s(ws2.cell(r2, ci), h, bg, _hfont(sz=10), _al(), _thin())
        ws2.column_dimensions[get_column_letter(ci)].width = 20 if ci > 2 else (22 if ci == 1 else 10)
    ws2.row_dimensions[r2].height = 18
    r2 += 1

    for role in ROLES:
        loading = LOADING.get(role, False)
        _s(ws2.cell(r2, 1), role,
           C_ORANGE, _font("000000", 10, True), _al("right"), _thin())
        _s(ws2.cell(r2, 2), "☑" if loading else "☐",
           C_ORANGE if loading else C_LIGHT_GRAY,
           _font("000000", 11, loading), _al(), _thin())
        for ci in range(3, 3 + len(day_cols)):
            _s(ws2.cell(r2, ci), "", C_TPL_INPUT, _font(sz=10), _al(), _thin())
        ws2.row_dimensions[r2].height = 17
        r2 += 1

    # ── Sheet C: ROOMS INPUT ───────────────────────────────────────────────
    ws3 = wb.create_sheet("לינה — קלט")
    ws3.sheet_view.rightToLeft  = True
    ws3.sheet_view.showGridLines = False

    r3 = 1
    _merge(ws3, r3, 1, r3, 6,
           "לינה — מלא שמות לפי חדר ולילה",
           C_PURPLE, _hfont(sz=12), _medium())
    ws3.row_dimensions[r3].height = 24
    r3 += 1

    # headers
    night_hdrs = ["צימר / יחידה", "חדר", "לילה א'", "לילה ב'", "לילה ג'", "לילה ד'"]
    for ci, h in enumerate(night_hdrs, 1):
        bg = C_PURPLE if ci <= 2 else C_PURPLE
        ws3.column_dimensions[get_column_letter(ci)].width = 24 if ci > 2 else (20 if ci == 1 else 12)
        _s(ws3.cell(r3, ci), h, bg, _hfont(sz=10), _al(), _thin())
    ws3.row_dimensions[r3].height = 18
    r3 += 1

    sample_rooms = [
        ("דירה",    "חדר 1", "", "", "", ""),
        ("דירה",    "חדר 2", "", "", "", ""),
        ("דירה",    "חדר 3", "", "", "", ""),
        ("דירה",    "חדר 4", "", "", "", ""),
        ("זוהר",    "חדר 1", "", "", "", ""),
        ("זוהר",    "חדר 2", "", "", "", ""),
        ("זוהר",    "חדר 3", "", "", "", ""),
        ("בקתה 7",  "חדר 1", "", "", "", ""),
        ("בקתה 7",  "חדר 2", "", "", "", ""),
        ("בקתה 8",  "חדר 1", "", "", "", ""),
        ("בקתה 8",  "חדר 2", "", "", "", ""),
        ("שיזף",    "חדר 1", "", "", "", ""),
        ("שיזף",    "חדר 2", "", "", "", ""),
    ]
    prev_unit = None
    unit_start_row = r3
    for sr in sample_rooms:
        unit = sr[0]
        if unit != prev_unit:
            if prev_unit is not None:
                # merge unit column for previous block
                pass
            prev_unit = unit
        for ci, val in enumerate(sr, 1):
            bg = C_LAVENDER if ci <= 2 else C_TPL_INPUT
            ft = _font("3D2566", 10, ci <= 2) if ci <= 2 else _font(sz=10)
            _s(ws3.cell(r3, ci), val, bg, ft, _al(), _thin())
        ws3.row_dimensions[r3].height = 17
        r3 += 1

    wb.save(output_path)
    print(f"✓ Input template saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE RUN — reproduces week 1 from reference file
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # 1. Generate the blank input template
    generate_input_template(
        "/mnt/user-data/outputs/logistics-agent-he/weekly_input_template.xlsx"
    )

    # 2. Generate a sample formatted plan (week 1 data)
    sample = {
        "week_label": "שבוע 1  |  12.04 – 16.04.26",
        "experiments": [
            {
                "name": "עין יהב",
                "days": ["שני — ירידה", "שלישי", "רביעי", "חמישי"],
                "roles": [
                    {"role": "מנהל ניסוי",   "loading": False, "assignments": ["ספי",    "ספי",    "ספי",    "ספי"]},
                    {"role": "מטיס חוץ",     "loading": False, "assignments": ["רוי שיבר","רוי שיבר / ליאור זהבי","ליאור זהבי","ליאור זהבי"]},
                    {"role": "מטיס חוץ",     "loading": False, "assignments": ["",       "",       "אסף אלוש",""]},
                    {"role": "מטיס פנים",    "loading": False, "assignments": ["תמיר",   "תמיר",   "דן",     "דן"]},
                    {"role": "מטיס פנים",    "loading": False, "assignments": ["",       "",       "אסף אלוש",""]},
                    {"role": "מהנדס בקרה",   "loading": False, "assignments": ["",       "ארז",    "ארז",    ""]},
                    {"role": "מהנדס בקרה",   "loading": False, "assignments": ["",       "",       "",       ""]},
                    {"role": "מהנדס מוביל",  "loading": False, "assignments": ["",       "",       "",       ""]},
                    {"role": "מהנדס",        "loading": False, "assignments": ["",       "",       "",       ""]},
                    {"role": "טכנאי מטוסים", "loading": True,  "assignments": ["שי",    "שי",     "שי",     "שי"]},
                    {"role": "טכנאי מטוסים", "loading": True,  "assignments": ["אליאור","אליאור", "אליאור", "אליאור"]},
                    {"role": "טכנאי מטוסים", "loading": True,  "assignments": ["אסיף",  "אסיף",   "אסיף — בוקר",""]},
                    {"role": "בטיחות",       "loading": True,  "assignments": ["נטע",   "נטע",    "נטע",    "נטע"]},
                    {"role": "אופרציה",      "loading": True,  "assignments": ["קירה",  "קירה",   "קירה",   "קירה"]},
                    {"role": "אופרציה",      "loading": True,  "assignments": ["יהב",   "יהב",    "יהב",    "יהב"]},
                    {"role": "אופרציה",      "loading": False, "assignments": ["",       "",       "",       ""]},
                    {"role": "לוגיסטיקה",    "loading": True,  "assignments": ["תומר",  "תומר",   "תומר",   "תומר"]},
                    {"role": "לוגיסטיקה",    "loading": True,  "assignments": ["ג'ק",   "ג'ק",    "ג'ק",    ""]},
                    {"role": "לוגיסטיקה",    "loading": False, "assignments": ["",       "",       "",       ""]},
                    {"role": "מערך קרקעי",   "loading": False, "assignments": ["",       "",       "",       ""]},
                    {"role": "מערך קרקעי",   "loading": False, "assignments": ["",       "",       "",       ""]},
                ],
                "totals": [10, 14, 11, 11]
            },
            {
                "name": "מבוא חורון",
                "days": ["שני", "חמישי"],
                "roles": [
                    {"role": "מנהל ניסוי",   "loading": False, "assignments": ["",         ""]},
                    {"role": "מטיס חוץ",     "loading": False, "assignments": ["ליאור זהבי","רוי שיבר"]},
                    {"role": "מטיס חוץ",     "loading": False, "assignments": ["",         ""]},
                    {"role": "מטיס פנים",    "loading": False, "assignments": ["בר שוורץ", ""]},
                    {"role": "מטיס פנים",    "loading": False, "assignments": ["",         "אמיר אבני"]},
                    {"role": "מהנדס בקרה",   "loading": False, "assignments": ["ניב",      ""]},
                    {"role": "מהנדס בקרה",   "loading": False, "assignments": ["",         ""]},
                    {"role": "מהנדס מוביל",  "loading": False, "assignments": ["",         ""]},
                    {"role": "מהנדס",        "loading": False, "assignments": ["",         ""]},
                    {"role": "טכנאי מטוסים", "loading": True,  "assignments": ["אסיף",     ""]},
                    {"role": "טכנאי מטוסים", "loading": True,  "assignments": ["",         ""]},
                    {"role": "טכנאי מטוסים", "loading": True,  "assignments": ["",         ""]},
                    {"role": "בטיחות",       "loading": True,  "assignments": ["אריאל",    "אריאל"]},
                    {"role": "אופרציה",      "loading": True,  "assignments": ["אברהם",    "אברהם"]},
                    {"role": "אופרציה",      "loading": True,  "assignments": ["",         ""]},
                    {"role": "אופרציה",      "loading": False, "assignments": ["",         ""]},
                    {"role": "לוגיסטיקה",    "loading": True,  "assignments": ["",         ""]},
                    {"role": "לוגיסטיקה",    "loading": True,  "assignments": ["",         ""]},
                    {"role": "לוגיסטיקה",    "loading": False, "assignments": ["",         ""]},
                    {"role": "מערך קרקעי",   "loading": False, "assignments": ["",         ""]},
                    {"role": "מערך קרקעי",   "loading": False, "assignments": ["",         ""]},
                ],
                "totals": [8, 7]
            }
        ],
        "day_labels": ["יום שני", "יום שלישי", "יום רביעי", "יום חמישי"],
        "vehicles": [
            {
                "vehicle": "דוקאטו",
                "capacity": 5,
                "needs_trailer_license": False,
                "days": [
                    {"day": "יום שני",   "outbound": {"route": "תל אביב > עין יהב",    "commander": "שי",    "passengers": ["אליאור"]},              "return": None},
                    {"day": "יום שלישי", "outbound": None, "return": None},
                    {"day": "יום רביעי", "outbound": None, "return": None},
                    {"day": "יום חמישי", "outbound": None, "return": {"route": "עין יהב > תל אביב", "commander": "שי",    "passengers": ["אליאור"]}},
                ]
            },
            {
                "vehicle": "טנדר #1 + מתדלקת סולר + נגרר ערבי",
                "capacity": 4,
                "needs_trailer_license": True,
                "days": [
                    {"day": "יום שני",   "outbound": {"route": "תל אביב > עין יהב",    "commander": "תומר",  "passengers": ["ספי", "קירה"]},          "return": None},
                    {"day": "יום שלישי", "outbound": None, "return": None},
                    {"day": "יום רביעי", "outbound": None, "return": None},
                    {"day": "יום חמישי", "outbound": None, "return": {"route": "עין יהב > תל אביב", "commander": "תומר",  "passengers": ["ספי", "ליאור זהבי"]}},
                ]
            },
            {
                "vehicle": "טנדר #2 + מתדלקת דס\"ל - גדול",
                "capacity": 4,
                "needs_trailer_license": True,
                "days": [
                    {"day": "יום שני",   "outbound": {"route": "תל אביב > עין יהב",    "commander": "יהב",   "passengers": ["ג'ק", "ספי"]},           "return": None},
                    {"day": "יום שלישי", "outbound": None, "return": None},
                    {"day": "יום רביעי", "outbound": None, "return": None},
                    {"day": "יום חמישי", "outbound": None, "return": {"route": "עין יהב > תל אביב", "commander": "יהב",   "passengers": ["אסף אלוש", "ג'ק"]}},
                ]
            },
            {
                "vehicle": "טויוטה",
                "capacity": 5,
                "needs_trailer_license": False,
                "days": [
                    {"day": "יום שני",   "outbound": {"route": "תל אביב > מבוא חורון", "commander": "אדרי",  "passengers": ["אברהם"]},                "return": {"route": "מבוא חורון > תל אביב", "commander": "אדרי", "passengers": ["אברהם"]}},
                    {"day": "יום שלישי", "outbound": None, "return": None},
                    {"day": "יום רביעי", "outbound": None, "return": None},
                    {"day": "יום חמישי", "outbound": None, "return": None},
                ]
            },
        ],
        "accommodation": {
            "nights": ["לילה ב' (13.04)", "לילה ג' (14.04)", "לילה ד' (15.04)"],
            "night_totals": [10, 9, 10],
            "hostels": [
                {
                    "name": "רגע בערבה",
                    "units": [
                        {"name": "דירה",   "rooms": [
                            {"name": "חדר 1", "capacity": 4, "nights": [["תומר","ג'ק","יהב","שי"],   ["תומר","ג'ק","יהב"],          ["תומר","יהב","שי"],       []]},
                            {"name": "חדר 2", "capacity": 4, "nights": [[],[],[],[]]},
                            {"name": "חדר 3", "capacity": 4, "nights": [[],[],[],[]]},
                            {"name": "חדר 4", "capacity": 4, "nights": [[],[],[],[]]},
                        ]},
                        {"name": "זוהר",   "rooms": [
                            {"name": "חדר 1", "capacity": 3, "nights": [["תמיר","אליאור","רוי שיבר"],["אליאור","רוי שיבר","אסיף"],  ["אליאור","רוי שיבר","אסיף"],[]]},
                            {"name": "חדר 2", "capacity": 3, "nights": [[],[],[],[]]},
                            {"name": "חדר 3", "capacity": 3, "nights": [[],[],[],[]]},
                        ]},
                        {"name": "בקתה 7", "rooms": [
                            {"name": "חדר 1", "capacity": 3, "nights": [["ספי"],                     ["ספי"],                       ["ספי","ארז"],              []]},
                            {"name": "חדר 2", "capacity": 3, "nights": [[],[],[],[]]},
                        ]},
                        {"name": "בקתה 8", "rooms": [
                            {"name": "חדר 1", "capacity": 2, "nights": [["נטע","קירה"],               ["נטע","קירה"],                ["נטע","קירה"],             []]},
                            {"name": "חדר 2", "capacity": 2, "nights": [[],[],[],[]]},
                        ]},
                        {"name": "שיזף",   "rooms": [
                            {"name": "חדר 1", "capacity": 3, "nights": [[],                           ["ליאור זהבי","דן"],            [],                         []]},
                            {"name": "חדר 2", "capacity": 3, "nights": [[],[],[],[]]},
                        ]},
                    ]
                }
            ]
        }
    }

    generate_weekly_plan(
        sample,
        "/mnt/user-data/outputs/logistics-agent-he/weekly_plan_week1.xlsx"
    )
